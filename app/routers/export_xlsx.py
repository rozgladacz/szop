from __future__ import annotations

from io import BytesIO
from typing import Any

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import RedirectResponse, StreamingResponse
from openpyxl import Workbook
from sqlalchemy.orm import Session

from .. import models
from ..db import get_db
from ..security import get_current_user
from ..services import costs, utils
from .export import (
    _army_rule_labels,
    _army_spell_entries,
    _load_roster_for_export,
)
from .rosters import (
    _classification_map,
    _ensure_roster_view_access,
    _roster_unit_export_data,
    _roster_unit_loadout,
)


router = APIRouter(prefix="/export", tags=["export"])


def _abilities_text(passives: list[str], actives: list[str], auras: list[str]) -> str:
    parts: list[str] = []
    if passives:
        parts.append(f"Pasywne: {', '.join(passives)}")
    if actives:
        parts.append(f"Aktywne: {', '.join(actives)}")
    if auras:
        parts.append(f"Aury: {', '.join(auras)}")
    return "\n".join(parts) if parts else "-"


def _weapon_details_text(details: list[dict[str, Any]]) -> str:
    if not details:
        return "-"
    lines: list[str] = []
    for weapon in details:
        name = weapon.get("name") or "Broń"
        count = weapon.get("count") or 0
        range_value = weapon.get("range") or "-"
        attacks = weapon.get("attacks") or "-"
        ap_value = weapon.get("ap") if weapon.get("ap") is not None else "-"
        traits = weapon.get("traits") or "-"
        lines.append(
            f"{name} × {count} | Z: {range_value} | Ataki: {attacks} | AP: {ap_value} | Cechy: {traits}"
        )
    return "\n".join(lines)


def _append_roster_sheet(
    workbook: Workbook,
    entries: list[dict[str, Any]],
    roster_total: float,
    spells: list[dict[str, Any]] | None = None,
    army_rules: list[str] | None = None,
) -> None:
    sheet = workbook.active
    sheet.title = "Lista"
    header = [
        "Jednostka",
        "Oddział",
        "Ilość",
        "Jakość",
        "Obrona",
        "Wytrzymałość",
        "Zdolności",
        "Uzbrojenie",
        "Suma [pkt]",
    ]
    column_widths = [len(str(value)) for value in header]

    if army_rules:
        rule_text = ", ".join(army_rules)
        header_row = [f"Zasady armii: {rule_text}"]
        sheet.append(header_row)
        column_widths[0] = max(column_widths[0], len(header_row[0]))
        sheet.append([])

    sheet.append(header)

    for entry in entries:
        total_value = float(entry.get("total_cost", 0.0))
        rounded_value = entry.get("rounded_total_cost")
        if rounded_value is None:
            rounded_value = utils.round_points(total_value)
        row = [
            entry.get("unit_name"),
            entry.get("custom_name") or "",
            entry.get("count"),
            entry.get("quality"),
            entry.get("defense"),
            entry.get("toughness"),
            _abilities_text(
                entry.get("passive_labels", []),
                entry.get("active_labels", []),
                entry.get("aura_labels", []),
            ),
            _weapon_details_text(entry.get("weapon_details", [])),
            rounded_value,
        ]
        for index, value in enumerate(row):
            if index < len(column_widths):
                text = "" if value is None else str(value)
                column_widths[index] = max(column_widths[index], len(text))
        sheet.append(row)

    total_row = [
        "",
        "",
        "",
        "",
        "",
        "",
        "Razem",
        "",
        utils.round_points(roster_total),
    ]
    for index, value in enumerate(total_row):
        if index < len(column_widths):
            text = "" if value is None else str(value)
            column_widths[index] = max(column_widths[index], len(text))
    sheet.append(total_row)
    if spells:
        sheet.append([])
        spell_header = ["Koszt mocy", "Trudność", "Zaklęcie"]
        for index, value in enumerate(spell_header):
            text = "" if value is None else str(value)
            column_widths[index] = max(column_widths[index], len(text))
        sheet.append(spell_header)
        for spell in spells:
            difficulty = spell.get("difficulty")
            row = [
                spell.get("cost"),
                f"{difficulty}+" if difficulty is not None else "",
                spell.get("label"),
            ]
            for index, value in enumerate(row):
                text = "" if value is None else str(value)
                column_widths[index] = max(column_widths[index], len(text))
            sheet.append(row)
    for index, max_length in enumerate(column_widths, start=1):
        column_letter = sheet.cell(row=1, column=index).column_letter
        sheet.column_dimensions[column_letter].width = min(max_length + 2, 60)


def _append_weapons_sheet(
    workbook: Workbook, entries: list[dict[str, Any]], army_rules: list[str] | None = None
) -> None:
    sheet = workbook.create_sheet("Zbrojownia")
    header = ["Nazwa", "Ilość", "Zasięg", "Ataki", "AP", "Cechy"]
    column_widths = [len(str(value)) for value in header]

    if army_rules:
        rule_header = [f"Zasady armii: {', '.join(army_rules)}"]
        sheet.append(rule_header)
        column_widths[0] = max(column_widths[0], len(rule_header[0]))
        sheet.append([])

    sheet.append(header)
    aggregated: dict[tuple[str, str, str, str, str], int] = {}
    for entry in entries:
        for weapon in entry.get("weapon_details", []):
            name = weapon.get("name") or "Broń"
            range_value = weapon.get("range") or "-"
            attacks = str(weapon.get("attacks") or "-")
            ap_value = (
                str(weapon.get("ap")) if weapon.get("ap") is not None else "-"
            )
            traits = weapon.get("traits") or "-"
            key = (name, str(range_value), attacks, ap_value, traits)
            aggregated[key] = aggregated.get(key, 0) + int(weapon.get("count") or 0)
    for (name, range_value, attacks, ap_value, traits), count in sorted(aggregated.items()):
        row = [name, count, range_value, attacks, ap_value, traits]
        for index, value in enumerate(row):
            text = "" if value is None else str(value)
            column_widths[index] = max(column_widths[index], len(text))
        sheet.append(row)
    for index, max_length in enumerate(column_widths, start=1):
        column_letter = sheet.cell(row=1, column=index).column_letter
        sheet.column_dimensions[column_letter].width = min(max_length + 2, 50)


@router.get("/xlsx/{roster_id}")
def export_xlsx(
    roster_id: int,
    db: Session = Depends(get_db),
    current_user: models.User | None = Depends(get_current_user(optional=True)),
):
    if current_user is None:
        return RedirectResponse(url="/auth/login", status_code=303)
    roster = _load_roster_for_export(db, roster_id)
    if not roster:
        raise HTTPException(status_code=404)
    _ensure_roster_view_access(roster, current_user)

    total_cost, _ = costs.recalculate_roster_costs(roster)
    workbook = Workbook()
    unit_cache: dict[int, dict[str, Any]] = {}
    loadouts: dict[int, dict[str, Any]] = {}
    for roster_unit in roster.roster_units:
        unit_id = getattr(roster_unit, "id", None)
        if unit_id is None:
            continue
        loadouts[unit_id] = _roster_unit_loadout(roster_unit)
    classifications, totals_by_id = _classification_map(roster.roster_units, loadouts)
    entries = []
    for roster_unit in roster.roster_units:
        unit_id = getattr(roster_unit, "id", None)
        entries.append(
            _roster_unit_export_data(
                roster_unit,
                unit_cache=unit_cache,
                loadout_override=loadouts.get(unit_id),
                classification=classifications.get(unit_id),
                totals=totals_by_id.get(unit_id),
            )
        )
    spell_entries = _army_spell_entries(roster, entries)
    army_rules = _army_rule_labels(getattr(roster, "army", None))
    _append_roster_sheet(
        workbook, entries, total_cost, spell_entries, army_rules=army_rules
    )
    _append_weapons_sheet(workbook, entries, army_rules=army_rules)

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    filename = f"roster_{roster_id}_{utils.round_points(total_cost)}.xlsx"
    headers = {"Content-Disposition": f"attachment; filename={filename}"}
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )
